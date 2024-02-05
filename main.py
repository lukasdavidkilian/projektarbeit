from openpyxl import Workbook
import time

from generate_meal_plans import generate_meal_plans
from optimized_generate_meal_plans import optimized_generate_meal_plans

from food_sets import nutrition_set_1, nutrition_set_2, nutrition_set_3, nutrition_set_4, nutrition_set_5

food_lists = {
    "0": [],
    "1": nutrition_set_1,
    "2": nutrition_set_2,
    "3": nutrition_set_3,
    "4": nutrition_set_4,
    "5": nutrition_set_5
}

goal_sets = [
    [160, 80, 140],
    [200, 75, 220],
    [250, 65, 300],
    [300, 100, 400],
    [1, 75, 220],
    [200, 1, 220],
    [200, 75, -1],
]


def test_algorithm(food_set, goal, algorithm, test_runs=100):
    total_duration = 0
    success_count = 0
    total_counter = 0
    for _ in range(test_runs):
        start_time = time.time()
        results = algorithm(food_set, goal)
        duration = (time.time() - start_time) * 1000
        total_duration += duration
        if results[0] is not None:
            success_count += 1
        total_counter += results[1]
        print(_)
    average_duration = total_duration / test_runs
    success_rate = (success_count / test_runs) * 100
    average_counter = total_counter / test_runs
    return average_duration, average_counter, success_rate


def main():
    workbook = Workbook()
    sheet1 = workbook.create_sheet(title="Ergebnisse Algo 1")
    sheet2 = workbook.create_sheet(title="Ergebnisse Algo 2")

    headers = [
        "Food Set", "Anzahl Lebensmittel in der Liste", "Protein Ziel", "Fett Ziel",
        "Kohlenhydrate Ziel", "Durchschnittliche Dauer (ms)", "Durchschnittlicher Counter", "Plan gefunden (%)"
    ]
    sheet1.append(headers)
    sheet2.append(headers)

    for food_set_name, food_set in food_lists.items():
        food_set_count = len(food_set)
        for goal in goal_sets:
            duration_1, counter_1, success_rate_1 = test_algorithm(food_set, goal, generate_meal_plans)
            row = [food_set_name, food_set_count, goal[0], goal[1], goal[2], duration_1, counter_1, success_rate_1]
            sheet1.append(row)

            duration_2, counter_2, success_rate_2 = test_algorithm(food_set, goal, optimized_generate_meal_plans)
            row = [food_set_name, food_set_count, goal[0], goal[1], goal[2], duration_2, counter_2, success_rate_2]
            sheet2.append(row)

    del workbook['Sheet']
    workbook.save("Ergebnisse.xlsx")


if __name__ == "__main__":
    main()
